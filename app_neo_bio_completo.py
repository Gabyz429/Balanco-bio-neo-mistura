
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Balanço Neo Bio (Completo)", layout="wide")

# --------- THEME / STYLE ---------
PRIMARY = "#0B7A75"
NEG     = "#C62828"
POS     = "#2E7D32"
BG_APP  = "#F6F7F9"
BG_CARD = "#FFFFFF"

st.markdown(f"""
<style>
.stApp, section.main, div[data-testid="stAppViewContainer"], div.block-container {{
  background: {BG_APP} !important;
}}
.badge {{
  display:inline-block; padding:2px 8px; border-radius:999px;
  font-size:.75rem; border:1px solid #e5e7eb; background:#fff; margin-left:6px;
}}
.badge.formula {{ background:#E8F5E9; border-color:#C8E6C9; }}
.badge.variar  {{ background:#FFF3E0; border-color:#FFE0B2; }}
.badge.fixo    {{ background:#ECEFF1; border-color:#CFD8DC; }}
.card {{
  background:{BG_CARD}; border:1px solid #eef0f2; border-radius:16px; padding:14px 16px; box-shadow:0 1px 3px rgba(0,0,0,.05);
}}
.grid2 {{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
.grid3 {{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}}
hr{{margin:12px 0 18px 0}}
h2{{margin-top:6px}}
</style>
""", unsafe_allow_html=True)

st.title("Balanço Neo Bio — Aba Completo")

# --------- Load defaults from workbook ---------
@st.cache_data
def load_defaults(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb["Completo"]
    # Map needed cells (row, col) -> value
    def v(r,c): return ws.cell(r,c).value

    defaults = {
        # Dados da Bio (coluna C)
        "C4_Cana": v(4,3) or 0.0,
        "C5_K_cana": v(5,3) or 0.0,
        "C6_vazao_vinho": v(6,3) or 0.0,
        "C8_ds": v(8,3) or 0.0,
        "C9_gl": v(9,3) or 0.0,
        # Volante Neo - Vinho
        "C19_vazao": v(19,3) or 0.0,
        "C20_ds": v(20,3) or 0.0,
        "C21_gl": v(21,3) or 0.0,
        # Mistura
        "H8_consumo_to_be": v(8,8) or 0.0,
    }
    return defaults

defaults = load_defaults("/mnt/data/Balanço Neo Bio (1).xlsx")

# --------- Helpers ---------
def kpi(label, value, aux=None):
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown(f"**{label}**")
    st.markdown(f"<div style='font-size:1.3rem;font-weight:800'>{value}</div>", unsafe_allow_html=True)
    if aux:
        st.caption(aux)
    st.markdown('</div>', unsafe_allow_html=True)

def fmt(x, nd=3):
    try:
        return f"{float(x):,.{nd}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(x)

# --------- Part 1: Dados da Bio ---------
st.header("1) Dados da Bio")
col1, col2, col3, col4 = st.columns(4)
with col1:
    C4 = defaults["C4_Cana"]; st.write(f"**Cana**: {fmt(C4)}"); st.markdown('<span class="badge fixo">Fixo</span>', unsafe_allow_html=True)
with col2:
    C5 = defaults["C5_K_cana"]; st.write(f"**K Cana (kg/t)**: {fmt(C5)}"); st.markdown('<span class="badge fixo">Fixo</span>', unsafe_allow_html=True)
with col3:
    C6 = st.number_input("Vazão vinho (m³/h) — C6", min_value=0.0, value=float(defaults["C6_vazao_vinho"]), step=1.0, format="%.3f"); st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)
with col4:
    st.empty()

col5, col6, col7 = st.columns(3)
with col5:
    C8 = st.number_input("%Ds — C8", min_value=0.0, value=float(defaults["C8_ds"]), step=0.1, format="%.3f"); st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)
with col6:
    C9 = st.number_input("Conc. GL — C9", min_value=0.0, value=float(defaults["C9_gl"]), step=0.1, format="%.3f"); st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)
with col7:
    st.empty()

# Fórmulas coluna C
C7  = (C4*C5/C6) if C6 else 0.0                                  # =C4*C5/C6
C10 = (-0.244*C9 + 4.564)                                         # =-0.244*C9 + 4.564
C11 = (C6*C9/96.0*C10)                                            # =C6*C9/96*C10
C12 = (C6*C9/96.0)                                                # =C6*C9/96
C13 = (C12*24.0)                                                  # =C12*24
C14 = (C12*1.2)                                                   # =C12*1.2
C15 = (C6 - C12*0.789 + C11 - C14)                                # =C6-C12*0.789+C11-C14
C16 = (C6 / C8 / C15) if (C8 and C15) else 0.0                    # =C6/C8/C15
C17 = (C5*C4/C15) if C15 else 0.0                                 # =C5*C4/C15

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("K vinho — C7", f"{fmt(C7)}", "Fórmula =C4*C5/C6")
kpi("Consumo específico (as is) — C10", f"{fmt(C10)}", "Fórmula =-0.244*C9 + 4.564")
kpi("V1 total as is — C11 (m³/h)", f"{fmt(C11)}", "Fórmula =C6*C9/96*C10")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("Etanol as is — C12 (m³/h)", f"{fmt(C12)}", "Fórmula =C6*C9/96")
kpi("Etanol to be — C13 (m³/dia)", f"{fmt(C13)}", "Fórmula =C12*24")
kpi("Flegmassa — C14 (m³/h)", f"{fmt(C14)}", "Fórmula =C12*1.2")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("Vinhaça — C15 (m³/h)", f"{fmt(C15)}", "Fórmula =C6 - C12*0.789 + C11 - C14")
kpi("Sólidos na vinhaça — C16 (%)", f"{fmt(C16*100,2)} %", "Fórmula =C6/C8/C15")
kpi("K vinhaça — C17", f"{fmt(C17)}", "Fórmula =C5*C4/C15")
st.markdown('</div>', unsafe_allow_html=True)

st.divider()

# --------- Part 2: Dados Neo ---------
st.header("2) Dados Neo (Volante Neo - Vinho)")
colN1, colN2, colN3 = st.columns(3)
with colN1:
    C19 = st.number_input("Vazão (m³/h) — C19", min_value=0.0, value=float(defaults["C19_vazao"]), step=1.0, format="%.3f"); st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)
with colN2:
    C20 = st.number_input("%Ds — C20", min_value=0.0, value=float(defaults["C20_ds"]), step=0.1, format="%.3f"); st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)
with colN3:
    C21 = st.number_input("Conc. GL — C21", min_value=0.0, value=float(defaults["C21_gl"]), step=0.1, format="%.3f"); st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)

st.divider()

# --------- Part 3: Dados da Mistura ---------
st.header("3) Dados da Mistura (H*)")

# H5, H6, H7 — Fórmulas
H5 = C19 + C6                                           # =C19+C6
H6 = ((C19*C20 + C6*C8) / H5) if H5 else 0.0            # =(C19*C20+C6*C8)/H5
H7 = ((C19*C21 + C6*C9) / H5) if H5 else 0.0            # =(C19*C21+C6*C9)/H5

# H8 — Variar
H8 = st.number_input("Consumo específico - to be — H8 (kg/L)", min_value=0.0, value=float(defaults["H8_consumo_to_be"]), step=0.1, format="%.3f")
st.markdown('<span class="badge variar">Variar</span>', unsafe_allow_html=True)

# H9..H17 — Fórmulas
H9  = H5*H7/96.0*H8                                     # =H5*H7/96*H8
H10 = H9 - C11                                          # =H9-C11 (Diferença V1)
H11 = H5*H7/96.0                                        # =H5*H7/96 (Etanol hidratado m³/h)
H12 = H11*24.0                                          # =H11*24 (Etanol dia)
H13 = H11*1.2                                           # =H11*1.2 (Flegmaça)
H14 = H5 - H11*0.786 + H9 - H13                         # =H5-H11*0.786+H9-H13
H15 = (H5*H6/H14) if H14 else 0.0                       # =H5*H6/H14 (Sólidos na vinhaça - to be)
H16 = H14 - C15                                         # =H14-C15 (Diferença de vinhaça)
H17 = (C4*C5/H14) if H14 else 0.0                       # =C4*C5/H14 (K vinhaça)

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("Vazão Mistura — H5 (m³/h)", f"{fmt(H5)}", "Fórmula =C19+C6")
kpi("%Ds Mistura — H6 (%)", f"{fmt(H6,2)} %", "Fórmula =(C19*C20+C6*C8)/H5")
kpi("Conc. w/w Mistura — H7 (%)", f"{fmt(H7,2)} %", "Fórmula =(C19*C21+C6*C9)/H5")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("V1 total to be — H9 (m³/h)", f"{fmt(H9)}", "Fórmula =H5*H7/96*H8")
kpi("Dif. V1 — H10 (m³/h)", f"{fmt(H10)}", "Fórmula =H9-C11")
kpi("Etanol hidratado — H11 (m³/h)", f"{fmt(H11)}", "Fórmula =H5*H7/96")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("Etanol dia — H12 (m³/dia)", f"{fmt(H12)}", "Fórmula =H11*24")
kpi("Flegmaça — H13 (m³/h)", f"{fmt(H13)}", "Fórmula =H11*1.2")
kpi("Vinhaça to be — H14 (m³/h)", f"{fmt(H14)}", "Fórmula =H5−H11*0.786+H9−H13")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="grid3">', unsafe_allow_html=True)
kpi("Sólidos na vinhaça to be — H15 (%)", f"{fmt(H15*100,2)} %", "Fórmula =H5*H6/H14")
kpi("Diferença de vinhaça — H16 (m³/h)", f"{fmt(H16)}", "Fórmula =H14−C15")
kpi("K vinhaça — H17", f"{fmt(H17)}", "Fórmula =C4*C5/H14")
st.markdown('</div>', unsafe_allow_html=True)

st.divider()

# --------- Part 4: Produções & Financeiro ---------
st.header("4) Produções & Financeiro")
with st.expander("Parâmetros financeiros (Variar)"):
    preco_etanol = st.number_input("Preço etanol (R$/m³)", min_value=0.0, value=2800.0, step=50.0, format="%.2f")
    preco_vinhaça = st.number_input("Preço/valor de vinhaça (R$/m³) — opcional", min_value=0.0, value=0.0, step=10.0, format="%.2f")

# Produções (usando 'as is' e 'to be')
producao_as_is_m3dia = C13
producao_to_be_m3dia = H12
delta_producao = producao_to_be_m3dia - producao_as_is_m3dia

receita_as_is = producao_as_is_m3dia * preco_etanol
receita_to_be = producao_to_be_m3dia * preco_etanol
delta_receita  = receita_to_be - receita_as_is

st.markdown('<div class="grid2">', unsafe_allow_html=True)
kpi("Etanol (as is) — C13 (m³/dia)", f"{fmt(producao_as_is_m3dia)}")
kpi("Etanol (to be) — H12 (m³/dia)", f"{fmt(producao_to_be_m3dia)}")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="grid2">', unsafe_allow_html=True)
kpi("Δ Produção (m³/dia)", f"{fmt(delta_producao)}")
kpi("Δ Receita (R$/dia)", f"{fmt(delta_receita,2)}")
st.markdown('</div>', unsafe_allow_html=True)

st.caption("Obs.: Em cada campo exibimos a observação: **Fixo**, **Variar** ou **Fórmula**, aplicando as fórmulas exatamente como na planilha.")
